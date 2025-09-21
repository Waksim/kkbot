# This file will contain the logic for exporting and importing card translations.
from io import BytesIO
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Any

from django.db import transaction
from django.db.models import QuerySet
from apps.cards.models import Card


def export_cards_to_excel(queryset: QuerySet[Card]) -> BytesIO:
    """
    Exports card data from a given queryset to an Excel workbook in memory.
    """
    workbook = openpyxl.Workbook()
    worksheet: Worksheet = workbook.active
    worksheet.title = 'Card Translations'

    # Define headers
    headers = [
        'card_id', 'name_en', 'name_ru', 'title_en', 'title_ru', 'description_en', 'description_ru'
    ]
    worksheet.append(headers)

    # Write data rows
    for card in queryset:
        row = [
            card.card_id,
            card.name_en,
            card.name_ru,
            card.title_en,
            card.title_ru,
            card.description_en,
            card.description_ru,
        ]
        worksheet.append(row)

    # Save workbook to a memory buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def get_translation_diff(file_stream: BytesIO) -> Dict[str, List[Dict[str, Any]]]:
    """
    Reads an Excel file from a stream and compares its content with the database.
    Returns a dictionary containing a diff of the changes and a list of errors.
    """
    diff = {'updates': [], 'errors': []}
    try:
        workbook = openpyxl.load_workbook(file_stream)
        worksheet = workbook.active

        # Get header row to map column names to indices
        header = [cell.value for cell in worksheet[1]]
        try:
            id_col = header.index('card_id')
            name_ru_col = header.index('name_ru')
            title_ru_col = header.index('title_ru')
            desc_ru_col = header.index('description_ru')
        except ValueError as e:
            diff['errors'].append(f"Missing required column in Excel file: {e}")
            return diff

        # Get all relevant cards from DB in one query
        card_ids = [row[id_col].value for row in worksheet.iter_rows(min_row=2) if row[id_col].value]
        cards_in_db = Card.objects.in_bulk(card_ids)

        for i, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            card_id = row[id_col].value
            if not card_id:
                continue

            card = cards_in_db.get(int(card_id))
            if not card:
                diff['errors'].append(f"Row {i}: Card with ID {card_id} not found in the database.")
                continue

            # Compare fields and record changes
            changes = []
            # Helper to compare and record a single field
            def _compare_field(field_name, new_value):
                old_value = getattr(card, field_name, '')
                # Treat None and empty strings as the same to avoid unnecessary diffs
                old_value = old_value or ''
                new_value = new_value or ''
                if old_value != new_value:
                    changes.append({
                        'field': field_name,
                        'old': old_value,
                        'new': new_value,
                    })

            _compare_field('name_ru', row[name_ru_col].value)
            _compare_field('title_ru', row[title_ru_col].value)
            _compare_field('description_ru', row[desc_ru_col].value)

            if changes:
                diff['updates'].append({
                    'card_id': card.card_id,
                    'name': card.name_en, # Use English name for display
                    'changes': changes
                })

    except Exception as e:
        diff['errors'].append(f"An unexpected error occurred: {e}")

    return diff


def apply_translations_from_excel(file_stream: BytesIO) -> List[str]:
    """
    Updates the database with the translations from an Excel file stream.
    Returns a list of error messages, if any.
    """
    errors = []
    try:
        workbook = openpyxl.load_workbook(file_stream)
        worksheet = workbook.active
        header = [cell.value for cell in worksheet[1]]

        # Column mapping
        id_col = header.index('card_id')
        name_ru_col = header.index('name_ru')
        title_ru_col = header.index('title_ru')
        desc_ru_col = header.index('description_ru')

        cards_to_update = []

        card_ids = [row[id_col].value for row in worksheet.iter_rows(min_row=2) if row[id_col].value]
        cards_in_db = Card.objects.in_bulk(card_ids)

        for i, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            card_id = row[id_col].value
            if not card_id:
                continue

            card = cards_in_db.get(int(card_id))
            if not card:
                errors.append(f"Row {i}: Card with ID {card_id} not found and was skipped.")
                continue

            # Update fields from Excel data
            card.name_ru = row[name_ru_col].value or ''
            card.title_ru = row[title_ru_col].value or ''
            card.description_ru = row[desc_ru_col].value or ''
            cards_to_update.append(card)

        # Perform a bulk update for efficiency
        with transaction.atomic():
            Card.objects.bulk_update(cards_to_update, ['name_ru', 'title_ru', 'description_ru'])

    except ValueError as e:
        errors.append(f"Missing required column in Excel file: {e}")
    except Exception as e:
        errors.append(f"An unexpected error occurred during the update: {e}")

    return errors
